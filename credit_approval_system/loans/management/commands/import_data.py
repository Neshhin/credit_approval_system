import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from loans.models import Customer, Loan
from datetime import datetime
from decimal import Decimal

class Command(BaseCommand):
    help = 'Import customer and loan data from Excel files'

    def handle(self, *args, **kwargs):
        self.stdout.write('Starting data import...')
        
        # Import customers
        self.import_customers()
        
        # Import loans
        self.import_loans()
        
        self.stdout.write(self.style.SUCCESS('Data import completed successfully!'))
    
    @transaction.atomic
    def import_customers(self):
        self.stdout.write('Importing customers...')
        
        wb = openpyxl.load_workbook('init_data/customer_data.xlsx')
        sheet = wb.active
        
        customers_created = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            customer_id, first_name, last_name, phone_number, monthly_salary, approved_limit, current_debt = row
            
            # Check if customer already exists
            if not Customer.objects.filter(customer_id=customer_id).exists():
                Customer.objects.create(
                    customer_id=customer_id,
                    first_name=first_name,
                    last_name=last_name,
                    age=30,  # Default age as it's not in the data
                    phone_number=str(phone_number),
                    monthly_salary=Decimal(str(monthly_salary)),
                    approved_limit=Decimal(str(approved_limit)),
                    current_debt=Decimal(str(current_debt)) if current_debt else Decimal('0')
                )
                customers_created += 1
        
        self.stdout.write(self.style.SUCCESS(f'Created {customers_created} customers'))
    
    @transaction.atomic
    def import_loans(self):
        self.stdout.write('Importing loans...')
        
        wb = openpyxl.load_workbook('init_data/loan_data.xlsx')
        sheet = wb.active
        
        loans_created = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            customer_id, loan_id, loan_amount, tenure, interest_rate, monthly_repayment, emis_paid_on_time, start_date, end_date = row
            
            try:
                customer = Customer.objects.get(customer_id=customer_id)
                
                # Check if loan already exists
                if not Loan.objects.filter(loan_id=loan_id).exists():
                    # Parse dates
                    if isinstance(start_date, str):
                        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                    if isinstance(end_date, str):
                        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                    
                    Loan.objects.create(
                        loan_id=loan_id,
                        customer=customer,
                        loan_amount=Decimal(str(loan_amount)),
                        tenure=int(tenure),
                        interest_rate=Decimal(str(interest_rate)),
                        monthly_repayment=Decimal(str(monthly_repayment)),
                        emis_paid_on_time=int(emis_paid_on_time),
                        start_date=start_date,
                        end_date=end_date,
                        is_active=(datetime.now().date() <= end_date)
                    )
                    loans_created += 1
            except Customer.DoesNotExist:
                self.stdout.write(self.style.WARNING(f'Customer {customer_id} not found for loan {loan_id}'))
                continue
        
        self.stdout.write(self.style.SUCCESS(f'Created {loans_created} loans'))
